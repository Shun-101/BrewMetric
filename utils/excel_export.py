"""
BrewMetric Excel Export Module
CSV and Excel export functionality for reports.
"""

import csv
import io
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from config import EXPORT_DATE_FORMAT, EXPORT_DECIMAL_PLACES


class ExcelExporter:
    """Export inventory and report data to CSV and Excel formats."""
    
    @staticmethod
    def export_inventory_csv(items: list, filepath: str = None) -> str:
        """
        Export inventory items to CSV.
        
        Args:
            items: List of InventoryItem objects
            filepath: Optional file path to save to
            
        Returns:
            CSV content as string or filepath if saved
        """
        output = io.StringIO()
        writer = csv.writer(output, quoting=csv.QUOTE_ALL)
        
        # Header
        headers = [
            "Name",
            "Category",
            "Current Stock",
            "Unit",
            "Min Threshold",
            "Unit Cost",
            "Total Value",
            "Expiration Date",
            "Location",
            "Status",
            "Last Updated"
        ]
        writer.writerow(headers)
        
        # Data rows
        for item in items:
            status = "Expired" if item.is_expired() else \
                    "Expiring Soon" if item.is_expiring_soon() else \
                    "Low Stock" if item.is_below_threshold() else \
                    "Healthy"
            
            writer.writerow([
                item.name,
                item.category,
                f"{item.quantity:.{EXPORT_DECIMAL_PLACES}f}",
                item.unit,
                f"{item.min_threshold:.{EXPORT_DECIMAL_PLACES}f}",
                f"${item.unit_cost:.{EXPORT_DECIMAL_PLACES}f}",
                f"${item.get_stock_value():.{EXPORT_DECIMAL_PLACES}f}",
                item.expiration_date.strftime(EXPORT_DATE_FORMAT) if item.expiration_date else "N/A",
                item.location or "N/A",
                status,
                item.updated_at.strftime(EXPORT_DATE_FORMAT)
            ])
        
        csv_content = output.getvalue()
        
        # Save to file if path provided
        if filepath:
            Path(filepath).parent.mkdir(parents=True, exist_ok=True)
            with open(filepath, 'w', encoding='utf-8') as f:
                f.write(csv_content)
            return filepath
        
        return csv_content
    
    @staticmethod
    def export_inventory_excel(items: list, filepath: str) -> str:
        """
        Export inventory items to Excel with formatting.
        
        Args:
            items: List of InventoryItem objects
            filepath: File path to save to
            
        Returns:
            Filepath
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventory"
        
        # Define styles
        header_fill = PatternFill(start_color="00A86B", end_color="00A86B", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        low_stock_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
        expiring_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        
        # Headers
        headers = [
            "Name",
            "Category",
            "Current Stock",
            "Unit",
            "Min Threshold",
            "Unit Cost",
            "Total Value",
            "Expiration Date",
            "Location",
            "Status",
            "Last Updated"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Data rows
        for row, item in enumerate(items, 2):
            status = "Expired" if item.is_expired() else \
                    "Expiring Soon" if item.is_expiring_soon() else \
                    "Low Stock" if item.is_below_threshold() else \
                    "Healthy"
            
            row_data = [
                item.name,
                item.category,
                f"{item.quantity:.{EXPORT_DECIMAL_PLACES}f}",
                item.unit,
                f"{item.min_threshold:.{EXPORT_DECIMAL_PLACES}f}",
                f"${item.unit_cost:.{EXPORT_DECIMAL_PLACES}f}",
                f"${item.get_stock_value():.{EXPORT_DECIMAL_PLACES}f}",
                item.expiration_date.strftime(EXPORT_DATE_FORMAT) if item.expiration_date else "N/A",
                item.location or "N/A",
                status,
                item.updated_at.strftime(EXPORT_DATE_FORMAT)
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.border = border
                
                # Apply color based on status
                if col == 10:  # Status column
                    if status == "Low Stock":
                        cell.fill = low_stock_fill
                    elif status == "Expiring Soon":
                        cell.fill = expiring_fill
                
                # Align numbers right
                if col in [3, 5, 6, 7]:
                    cell.alignment = Alignment(horizontal="right")
        
        # Adjust column widths
        column_widths = [25, 15, 15, 10, 15, 12, 15, 20, 15, 15, 20]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col)].width = width
        
        # Save file
        Path(filepath).parent.mkdir(parents=True, exist_ok=True)
        wb.save(filepath)
        
        return filepath
    
    @staticmethod
    def export_waste_log_csv(waste_logs: list, filepath: str = None) -> str:
        """
        Export waste log to CSV.
        
        Args:
            waste_logs: List of WasteLog objects
            filepath: Optional file path to save to
            
        Returns:
            CSV content as string or filepath if saved
        """
        output = io.StringIO()
        writer = csv.writer(output, quoting=csv.QUOTE_ALL)
        
        # Header
        headers = ["Date", "Item Name", "Quantity", "Unit", "Reason", "User", "Notes"]
        writer.writerow(headers)
        
        # Data rows
        for log in waste_logs:
            writer.writerow([
                log.created_at.strftime(EXPORT_DATE_FORMAT),
                log.inventory_item.name,
                f"{log.quantity:.{EXPORT_DECIMAL_PLACES}f}",
                log.inventory_item.unit,
                log.reason,
                log.user.full_name,
                log.notes or ""
            ])
        
        csv_content = output.getvalue()
        
        if filepath:
            Path(filepath).parent.mkdir(parents=True, exist_ok=True)
            with open(filepath, 'w', encoding='utf-8') as f:
                f.write(csv_content)
            return filepath
        
        return csv_content
    
    @staticmethod
    def export_audit_trail_csv(audit_trails: list, filepath: str = None) -> str:
        """
        Export audit trail to CSV.
        
        Args:
            audit_trails: List of AuditTrail objects
            filepath: Optional file path to save to
            
        Returns:
            CSV content as string or filepath if saved
        """
        output = io.StringIO()
        writer = csv.writer(output, quoting=csv.QUOTE_ALL)
        
        # Header
        headers = [
            "Timestamp",
            "User",
            "Action",
            "Entity Type",
            "Entity ID",
            "Description",
            "Old Values",
            "New Values",
            "IP Address"
        ]
        writer.writerow(headers)
        
        # Data rows
        for trail in audit_trails:
            writer.writerow([
                trail.created_at.strftime(EXPORT_DATE_FORMAT),
                trail.user.full_name,
                trail.action,
                trail.entity_type,
                trail.entity_id,
                trail.description or "",
                trail.old_values or "",
                trail.new_values or "",
                trail.ip_address or "N/A"
            ])
        
        csv_content = output.getvalue()
        
        if filepath:
            Path(filepath).parent.mkdir(parents=True, exist_ok=True)
            with open(filepath, 'w', encoding='utf-8') as f:
                f.write(csv_content)
            return filepath
        
        return csv_content
    
    @staticmethod
    def generate_filename(prefix: str = "export") -> str:
        """
        Generate a unique filename with timestamp.
        
        Args:
            prefix: Filename prefix
            
        Returns:
            Filename with timestamp
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return f"{prefix}_{timestamp}"
